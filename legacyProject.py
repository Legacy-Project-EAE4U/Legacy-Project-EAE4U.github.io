import barcode
from openpyxl import load_workbook

'''
#load excel file
workbook = load_workbook(filename="Legacy.xlsx")
 
#open workbook
sheet = workbook.active

print(sheet.max_row)
#modify the desired cell
sheet["A1"] = "Full Name"
 
#save the file
#workbook.save(filename="csv/output.xlsx")
'''

def createId(rowNumber):
    #make sure it's 8 numbers long
    string = list(str(rowNumber))
    for i in range(8 - len(string)):
        string.append("0") 

    #turn it back into a string
    string = str(string)

    #encode the string
    encoded = string.encode('utf-8')
    
    return(int.from_bytes(encoded, "little"))

print(createId(3234234))
